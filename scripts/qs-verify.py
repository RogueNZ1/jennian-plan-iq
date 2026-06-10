"""Dump the STRUCTURE of the live QS master's '5. Data Input House' sheet for
verification against the shipped export mapping. Labels, formulas, and template
dims ONLY — never rate/price columns (G, I) and no client data."""
import sys
from openpyxl import load_workbook

wb = load_workbook(sys.argv[1], read_only=True, data_only=False)
name = next((s for s in wb.sheetnames if "data input house" in s.lower()), None)
print(f"sheet: {name!r} | all sheets: {len(wb.sheetnames)}")
ws = wb[name]
print("── windows block A38:F74 (labels + template dims; G/I rates excluded) ──")
for r in range(38, 75):
    vals = [f"{c}{r}={ws[f'{c}{r}'].value!r}" for c in "ABCDEF" if ws[f"{c}{r}"].value is not None]
    if vals: print(" | ".join(vals))
print("── garage block rows 173-182 (labels + H counts only) ──")
for r in range(173, 183):
    vals = [f"{c}{r}={ws[f'{c}{r}'].value!r}" for c in "ABCDEFH" if ws[f"{c}{r}"].value is not None]
    if vals: print(" | ".join(vals))
print("── interior doors rows 184-196 (labels + H counts only) ──")
for r in range(184, 197):
    vals = [f"{c}{r}={ws[f'{c}{r}'].value!r}" for c in "ABCDEFH" if ws[f"{c}{r}"].value is not None]
    if vals: print(" | ".join(vals))

# ── FULL IQ IMPORT TAB + every Data-Input-House formula referencing it ──
iq = next((s for s in wb.sheetnames if "iq import" in s.lower()), None)
print(f"\n══ IQ IMPORT TAB: {iq!r} ══")
if iq:
    wi = wb[iq]
    for r in range(1, 51):
        vals = [f"{c}{r}={wi[f'{c}{r}'].value!r}" for c in "ABCDEF" if wi[f"{c}{r}"].value is not None]
        if vals: print(" | ".join(vals))
print("\n══ ALL 'IQ Import' FORMULA REFS IN DATA INPUT HOUSE (rows 1-260) ══")
for r in range(1, 261):
    for c in "ABCDEFGHIJ":
        v = ws[f"{c}{r}"].value
        if isinstance(v, str) and "IQ Import" in v:
            print(f"{c}{r} = {v}")
